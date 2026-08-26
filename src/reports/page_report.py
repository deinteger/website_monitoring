from __future__ import annotations
from datetime import datetime
from pathlib import Path
import os, re, shutil, tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SHEETS = ["\u81ea\u5408\u73b0\u51b5", "\u5185\u5bb9\u6700\u65b0\u6027", "\u94fe\u63a5\u00b7\u56fe\u50cf\u00b7\u9644\u4ef6", "HTML\u00b7\u65e0\u969c\u788d", "\u8bbf\u95ee\u72b6\u6001\u00b7\u6027\u80fd", "\u68c0\u67e5\u8303\u56f4\u00b7\u9057\u6f0f\u00b7\u5931\u8d25"]
# Correct Korean sheet names using escapes, avoiding source encoding corruption.
SHEETS = ["\uc885\ud569\ud604\ud669", "\ucf58\ud150\uce20 \ucd5c\uc2e0\uc131", "\ub9c1\ud06c\u00b7\uc774\ubbf8\uc9c0\u00b7\ucca8\ubd80", "HTML\u00b7\uc811\uadfc\uc131", "\uc811\uc18d\uc0c1\ud0dc\u00b7\uc131\ub2a5", "\uc810\uac80\ubc94\uc704\u00b7\ub204\ub77d\u00b7\uc2e4\ud328"]
VERDICTS = ["\uc815\uc0c1", "\uac80\ud1a0 \ud544\uc694", "\uc624\ub958", "\uc810\uac80 \ubd88\uac00", "\uc81c\uc678"]
COLORS = dict(zip(VERDICTS, ["C6EFCE","FFEB9C","FFC7CE","D9E1F2","E7E6E6"]))
CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def clean(value):
    value = "" if value is None else str(value)
    value = CONTROL.sub("", value)[:32767]
    return "'" + value if value.startswith(("=", "+", "-", "@")) else value

def safe_url(value):
    value = clean(value)
    return value if value.startswith(("https://", "http://")) else ""

class PageReportGenerator:
    def __init__(self, output_root="output", timezone_name="Asia/Seoul"):
        self.output_root = Path(output_root); self.timezone_name = timezone_name

    def build(self, payload=None, run_metadata=None):
        p = payload or {}; wb = Workbook(); ws = wb.active; ws.title = SHEETS[0]
        headers = ["\ub300\uc0c1 \uc0ac\uc774\ud2b8", "\uc804\uccb4 \ud398\uc774\uc9c0", *VERDICTS, "\uc2e0\uaddc", "\uc0ad\uc81c", "\ubcc0\uacbd", "\ud574\uacb0", "\uc218\uc9d1 \uc131\uacf5", "\ub204\ub77d\ubc94\uc704 \uc874\uc7ac"]
        self._write(ws, headers, [])
        ws.delete_rows(2)
        pages = list(p.get("page_results") or []); stats = p.get("site_stats") or {}
        targets = sorted({str(x.get("target_id", "")) for x in pages}) or ["\uc804\uccb4"]
        for target in targets:
            rows = [x for x in pages if target == "\uc804\uccb4" or str(x.get("target_id", "")) == target]
            counts = {v: sum(x.get("verdict") == v for x in rows) for v in VERDICTS}
            if target == "\uc804\uccb4" and stats.get("verdict_counts"): counts.update({v: stats["verdict_counts"].get(v, 0) for v in VERDICTS})
            changes = [sum(x.get("inventory_change_status") == v for x in rows) for v in ("\uc2e0\uaddc", "\uc0ad\uc81c", "\ubcc0\uacbd", "\ud574\uacb0")]
            ws.append([clean(target), len(rows), *[counts[v] for v in VERDICTS], *changes, "", "Y" if p.get("missing_scope") else "N"])
        details = [
            (SHEETS[1], ["target_id","menu_path","page_title","url","published_date","modified_date","extraction_method","problem_type","verdict","evidence","checked_at"], p.get("freshness_results", [])),
            (SHEETS[2], ["target_id","menu_path","source_page","source_url","resource_type","target_url","status_code","final_url","problem_type","severity","recommendation"], p.get("resource_results", [])),
            (SHEETS[3], ["target_id","menu_path","url","code","name","element","location","result","severity","evidence","recommendation","content_hash","cache_used","checked_at"], p.get("accessibility_issues", [])),
            (SHEETS[4], ["target_id","menu_path","url","status_code","final_url","redirect_count","total_seconds","received_bytes","console_error_count","ssl_result","mixed_content","verdict","reason","checked_at"], p.get("performance_results", [])),
            (SHEETS[5], ["target_id","source","url","normalized_url","classification","status","stage","failure_type","reason","retry_count","screenshot_path","checked_at","issue_key","lifecycle_status"], (p.get("coverage_records") or []) + (p.get("failure_details") or [])),
        ]
        for name, cols, records in details:
            sh = wb.create_sheet(name); self._write(sh, cols, records)
        for sh in wb.worksheets: self._format(sh)
        return wb

    def _write(self, ws, headers, records):
        ws.append(headers)
        records = list(records or [])
        for item in records or [{}]:
            vals = item if isinstance(item, (list, tuple)) else [item.get(h, "") for h in headers]
            ws.append([clean(v) for v in vals])
            for cell, header in zip(ws[ws.max_row], headers):
                target = safe_url(cell.value)
                if target and header in ("url", "source_url", "target_url"):
                    cell.hyperlink = target; cell.style = "Hyperlink"
                if header == "screenshot_path" and cell.value:
                    cell.hyperlink = str(cell.value); cell.style = "Hyperlink"
        if not records: ws.cell(ws.max_row, 1).value = "\uacb0\uacfc \uc5c6\uc74c"

    def _format(self, ws):
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F4E78"); c.alignment = Alignment(wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                if c.value in COLORS: c.fill = PatternFill("solid", fgColor=COLORS[c.value])
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for i, col in enumerate(ws.columns, 1): ws.column_dimensions[get_column_letter(i)].width = min(45, max(10, max((len(str(c.value or "")) for c in col), default=10) + 2))

    def save(self, payload, date=None, latest=True, run_metadata=None):
        date = date or datetime.now().strftime("%Y-%m-%d")
        dated = self.output_root / date / f"\ud398\uc774\uc9c0\ubcc4_\uc810\uac80\uacb0\uacfc_{date}.xlsx"; dated.parent.mkdir(parents=True, exist_ok=True)
        fd, tmp = tempfile.mkstemp(prefix=".page-report-", suffix=".xlsx", dir=str(dated.parent)); os.close(fd)
        try:
            self.build(payload, run_metadata).save(tmp); os.replace(tmp, dated)
            if latest:
                latest_path = self.output_root / "latest" / "\ud398\uc774\uc9c0\ubcc4_\uc810\uac80\uacb0\uacfc.xlsx"; latest_path.parent.mkdir(parents=True, exist_ok=True)
                tmp_latest = latest_path.with_suffix(".tmp.xlsx"); shutil.copyfile(dated, tmp_latest); os.replace(tmp_latest, latest_path)
        finally:
            if os.path.exists(tmp): os.unlink(tmp)
        return dated

def generate_page_report(payload, output_root="output", date=None): return PageReportGenerator(output_root).save(payload, date=date)
