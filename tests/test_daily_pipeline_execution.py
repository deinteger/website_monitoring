import json

from openpyxl import load_workbook

from src.daily_pipeline import DailyPipeline


def payload(**extra):
    value = {
        "page_results": [{"target_id": "fixture", "url": "https://fixture.test/a", "verdict": "정상"}],
        "coverage_summary": {"discovered_urls": 1},
    }
    value.update(extra)
    return value


def pipeline(tmp_path):
    return DailyPipeline(state_dir=tmp_path / "state", output_root=tmp_path / "output")


def test_normal_execution_returns_zero_and_creates_artifacts(tmp_path):
    result = pipeline(tmp_path).run_offline(payload(), run_id="normal", date="2026-08-26")
    assert result["exit_code"] == 0 and result["status"] == "completed"
    assert (tmp_path / "state" / "inventory.json").exists() and load_workbook(result["report_path"])


def test_partial_failure_returns_two_and_keeps_xlsx(tmp_path):
    result = pipeline(tmp_path).run_offline(payload(stage_failures={"checks": {"code": "timeout", "reason": "fixture timeout"}}), run_id="partial", date="2026-08-26")
    assert result["exit_code"] == 2 and result["status"] == "partial_failed"
    assert load_workbook(result["report_path"])


def test_partial_failure_still_generates_xlsx(tmp_path):
    result = pipeline(tmp_path).run_offline(payload(stage_failures={"inventory": "fixture failure"}), run_id="partial-xlsx", date="2026-08-26")
    assert result["exit_code"] == 2 and load_workbook(result["report_path"]).active.title


def test_invalid_configuration_returns_one(tmp_path):
    assert pipeline(tmp_path).run_offline([], run_id="bad")["exit_code"] == 1


def test_state_save_failure_returns_one(tmp_path, monkeypatch):
    p = pipeline(tmp_path)
    monkeypatch.setattr(p.state, "save_json", lambda *args, **kwargs: (_ for _ in ()).throw(OSError("disk unavailable")))
    assert p.run_offline(payload(), run_id="state-fail", date="2026-08-26")["exit_code"] == 1


def test_partial_failure_does_not_update_inventory_baseline(tmp_path):
    p = pipeline(tmp_path); p.state.save_json("inventory_baseline.json", {"fixture": {"version": "old"}})
    p.run_offline(payload(inventory_baseline={"fixture": {"version": "new"}}, stage_failures={"checks": "timeout"}), run_id="partial", date="2026-08-26")
    assert p.state.load_json("inventory_baseline.json")["fixture"]["version"] == "old"


def test_partial_failure_does_not_confirm_issue_resolution(tmp_path):
    p = pipeline(tmp_path); original = {"active_issues": [{"issue_key": "still-active", "lifecycle_status": "지속"}], "resolved_issues": []}
    p.state.save_json("issues.json", original)
    p.run_offline(payload(stage_failures={"aggregation": "incomplete"}), run_id="partial", date="2026-08-26")
    after = p.state.load_json("issues.json")
    assert after["active_issues"] == original["active_issues"] and not after["resolved_issues"]


def test_execution_stage_information_is_saved_with_matching_run_id(tmp_path):
    p = pipeline(tmp_path); result = p.run_offline(payload(), run_id="stage-run", date="2026-08-26")
    metadata = p.state.load_json("inventory.json")["run_metadata"]
    assert metadata["run_id"] == result["run_id"] == "stage-run"
    assert {"stage", "status", "started_at", "ended_at", "input_count", "processed_count", "failure_count", "failure_code", "failure_reason"} <= set(metadata["execution_stages"][0])


def test_stage_failure_code_and_reason_are_saved(tmp_path):
    p = pipeline(tmp_path); p.run_offline(payload(stage_failures={"checks": {"code": "fixture_timeout", "reason": "timeout without HTTP"}}), run_id="failed", date="2026-08-26")
    stage = next(x for x in p.state.load_json("inventory.json")["run_metadata"]["execution_stages"] if x["stage"] == "checks")
    assert stage["failure_code"] == "fixture_timeout" and stage["failure_reason"] == "timeout without HTTP"


def test_fixture_and_operational_transports_share_orchestration_order(tmp_path):
    class Transport:
        def __init__(self, data): self.data = data; self.calls = []
        def build_payload(self): self.calls.append("build_payload"); return self.data
    p = pipeline(tmp_path); fixture = Transport(payload()); p.run_with_transport(fixture, run_id="fixture", date="2026-08-26", transport_name="fixture"); fixture_trace = p.last_execution_trace[:]
    operational = Transport(payload()); p.run_with_transport(operational, run_id="operational", date="2026-08-27", transport_name="operational")
    assert fixture.calls == operational.calls == ["build_payload"]
    assert fixture_trace == p.last_execution_trace


def test_run_history_is_written_for_partial_result(tmp_path):
    result = pipeline(tmp_path).run_offline(payload(stage_failures={"checks": "timeout"}), run_id="history-run", date="2026-08-26")
    record = json.loads((tmp_path / "state" / "run_history.jsonl").read_text(encoding="utf-8").splitlines()[-1])
    assert record["run_id"] == result["run_id"] and record["exit_code"] == 2
