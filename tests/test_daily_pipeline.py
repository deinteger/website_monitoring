import json
from openpyxl import load_workbook
from src.daily_pipeline import DailyPipeline
from src.reports.page_report import SHEETS

def fixture():
    return {"page_results":[{"target_id":"nihhs","url":f"https://example.test/{i}","verdict":"정상" if i != 9 else "오류"} for i in range(10)], "inventory":{"nihhs":{"records":[]}}, "coverage_summary":{"discovered_urls":10}}

def test_offline_ten_page_pipeline_completes(tmp_path):
    result=DailyPipeline(state_dir=tmp_path/"state", output_root=tmp_path/"output").run_offline(fixture(),run_id="run-1",date="2026-08-25")
    assert result["status"]=="completed" and result["history"]["external_requests"]==0
    assert load_workbook(result["report_path"]).sheetnames == SHEETS

def test_offline_pipeline_writes_inventory_and_issues(tmp_path):
    DailyPipeline(state_dir=tmp_path/"state", output_root=tmp_path/"output").run_offline(fixture(),run_id="run-2",date="2026-08-25")
    assert (tmp_path/"state"/"inventory.json").exists() and (tmp_path/"state"/"issues.json").exists()

def test_offline_pipeline_writes_run_history(tmp_path):
    DailyPipeline(state_dir=tmp_path/"state", output_root=tmp_path/"output").run_offline(fixture(),run_id="run-3",date="2026-08-25")
    lines=(tmp_path/"state"/"run_history.jsonl").read_text(encoding="utf-8").splitlines(); assert json.loads(lines[-1])["run_id"]=="run-3"

def test_offline_pipeline_repeated_run_is_deterministic_rows(tmp_path):
    p=DailyPipeline(state_dir=tmp_path/"state", output_root=tmp_path/"output"); a=p.run_offline(fixture(),run_id="a",date="2026-08-25"); b=p.run_offline(fixture(),run_id="b",date="2026-08-26")
    wa=load_workbook(a["report_path"],data_only=True); wb=load_workbook(b["report_path"],data_only=True)
    assert [[c.value for c in r] for r in wa[SHEETS[0]].iter_rows()] == [[c.value for c in r] for r in wb[SHEETS[0]].iter_rows()]

def test_offline_pipeline_tolerates_missing_optional_payload(tmp_path):
    result=DailyPipeline(state_dir=tmp_path/"state", output_root=tmp_path/"output").run_offline({"page_results":[]},date="2026-08-25")
    assert result["status"]=="completed"

def test_raw_fixture_invokes_discovery_and_checks_without_http(tmp_path):
    base="https://fixture.test"
    fixture={"menu":{"main_selectors":["a"],"all_menu_selectors":["a"],"all_menu_paths":[],"sitemap_path":"/sitemap.xml"},"responses":{
        base+"/":{"status_code":200,"html":"<html lang='ko'><head><title>home</title></head><body><a href='/a'>A</a></body></html>"},
        base+"/sitemap.xml":{"status_code":200,"text":"<urlset xmlns='http://www.sitemaps.org/schemas/sitemap/0.9'><url><loc>https://fixture.test/a</loc></url></urlset>"},
        base+"/a":{"status_code":200,"html":"<html lang='ko'><head><title>a</title></head><body></body></html>"}}}
    out=DailyPipeline(state_dir=tmp_path/"state",output_root=tmp_path/"output").run_raw_fixture(fixture,date="2026-08-25")
    assert out["status"] == "completed" and out["inventory"].request_count <= 10

def raw_page(html):
    return {"menu":{"main_selectors":["a"],"all_menu_selectors":["a"],"all_menu_paths":[],"sitemap_path":"/sitemap.xml"},"responses":{
        "https://fixture.test/":{"status_code":200,"html":"<html><body><a href='/a'>A</a></body></html>"},
        "https://fixture.test/sitemap.xml":{"status_code":200,"text":"<urlset xmlns='http://www.sitemaps.org/schemas/sitemap/0.9'><url><loc>https://fixture.test/a</loc></url></urlset>"},
        "https://fixture.test/a":{"status_code":200,"html":html}}}

def test_three_runs_create_distinct_dates_and_history(tmp_path):
    p=DailyPipeline(state_dir=tmp_path/"s",output_root=tmp_path/"o")
    for d in ("2026-08-25","2026-08-26","2026-08-27"): p.run_raw_fixture(raw_page("<html><body><input></body></html>"),date=d)
    assert len(list((tmp_path/"o").glob("20*/*.xlsx"))) == 3
    assert len((tmp_path/"s"/"run_history.jsonl").read_text().splitlines()) == 3

def test_content_hash_cache_reuses_second_accessibility_check(tmp_path):
    p=DailyPipeline(state_dir=tmp_path/"s",output_root=tmp_path/"o"); f=raw_page("<html><body><input></body></html>")
    p.run_raw_fixture(f,date="2026-08-25"); first=p.last_metrics["accessibility_runs"]
    p.run_raw_fixture(f,date="2026-08-26"); assert first >= 1 and p.last_metrics["cache_hash_reuses"] >= 1

def test_content_change_invalidates_hash_cache(tmp_path):
    p=DailyPipeline(state_dir=tmp_path/"s",output_root=tmp_path/"o"); p.run_raw_fixture(raw_page("<html><body><input></body></html>"),date="2026-08-25")
    p.run_raw_fixture(raw_page("<html><body><input aria-label='x'></body></html>"),date="2026-08-26"); assert p.last_metrics["accessibility_runs"] >= 1

def test_attachment_cache_file_is_preserved_between_runs(tmp_path):
    p=DailyPipeline(state_dir=tmp_path/"s",output_root=tmp_path/"o"); f=raw_page("<html><body><a href='/doc.pdf'>doc</a></body></html>")
    p.run_raw_fixture(f,date="2026-08-25"); p.run_raw_fixture(f,date="2026-08-26"); assert (tmp_path/"s"/"attachment_cache.json").exists()

def test_latest_matches_third_run(tmp_path):
    p=DailyPipeline(state_dir=tmp_path/"s",output_root=tmp_path/"o"); p.run_raw_fixture(raw_page("<html><body>x</body></html>"),date="2026-08-27")
    dated=next((tmp_path/"o"/"2026-08-27").glob("*.xlsx")); assert dated.read_bytes()==(tmp_path/"o"/"latest"/"페이지별_점검결과.xlsx").read_bytes()

def test_run_ids_are_distinct(tmp_path):
    p=DailyPipeline(state_dir=tmp_path/"s",output_root=tmp_path/"o"); a=p.run_raw_fixture(raw_page("<html>x</html>"),run_id="one",date="2026-08-25"); b=p.run_raw_fixture(raw_page("<html>x</html>"),run_id="two",date="2026-08-26"); assert a["run_id"] != b["run_id"]
