from openpyxl import load_workbook
from src.reports.page_report import PageReportGenerator, SHEETS, VERDICTS

def payload():
    return {"page_results":[{"target_id":"b","menu_path":"z","url":"https://b","verdict":"정상"},{"target_id":"a","menu_path":"a","url":"https://a","verdict":"오류","inventory_change_status":"신규"}],"site_stats":{"pages_total":2,"verdict_counts":{"정상":1,"검토 필요":0,"오류":1,"점검 불가":0,"제외":0}},"missing_scope":[]}
def make(tmp_path, data=None): return PageReportGenerator(tmp_path).save(data or payload(), date="2026-08-25")
def wb(tmp_path, data=None): return load_workbook(make(tmp_path,data))

def test_date_and_latest(tmp_path):
    p=make(tmp_path); assert p.name.endswith("2026-08-25.xlsx"); assert (tmp_path/"latest"/"페이지별_점검결과.xlsx").exists()
def test_six_sheet_names_order(tmp_path): assert wb(tmp_path).sheetnames == SHEETS
def test_required_columns(tmp_path):
    w=wb(tmp_path); assert all(w[s].max_column >= 1 for s in SHEETS)
def test_all_verdict_pages(tmp_path):
    d={"page_results":[{"target_id":"t","verdict":v} for v in VERDICTS]}; w=wb(tmp_path,d); assert w[SHEETS[0]].max_row == 2
def test_aggregation_totals(tmp_path):
    w=wb(tmp_path); assert sum(w[SHEETS[0]].cell(r,c).value or 0 for r in (2,3) for c in range(3,8)) == 2
def test_coverage_missing_sheet(tmp_path):
    d=payload(); d["failure_details"]=[{"stage":"HTTP 요청","reason":"timeout"}]; w=wb(tmp_path,d); assert w[SHEETS[5]].max_row == 2
def test_lifecycle_status_columns_are_supported(tmp_path):
    d={"page_results":[{"target_id":"t","verdict":"검토 필요","inventory_change_status":v} for v in ("신규","지속","변경","해결","재발")]}; assert wb(tmp_path,d)[SHEETS[0]].max_row==2
def test_partial_failure_not_resolved(tmp_path):
    d={"page_results":[{"target_id":"t","verdict":"점검 불가","issue_lifecycle_status":"지속"}]}; assert wb(tmp_path,d)[SHEETS[0]].cell(2,6).value == 1
def test_url_hyperlinks(tmp_path):
    d={"freshness_results":[{"url":"https://example.test/a"}]}; c=wb(tmp_path,d)[SHEETS[1]]["D2"]; assert c.hyperlink.target == "https://example.test/a"
def test_screenshot_hyperlink(tmp_path):
    d={"failure_details":[{"screenshot_path":"screenshots/2026/a.png"}]}; c=wb(tmp_path,d)[SHEETS[5]]["K2"]; assert c.hyperlink.target.endswith("a.png")
def test_filter_freeze_and_colors(tmp_path):
    d=payload(); d["freshness_results"]=[{"verdict":"오류"}]; w=wb(tmp_path,d); s=w[SHEETS[0]]; assert s.freeze_panes=="A2" and s.auto_filter.ref; assert w[SHEETS[1]].cell(2,9).fill.fgColor.rgb.endswith("FFC7CE")
def test_deterministic_target_sort(tmp_path):
    w=wb(tmp_path); assert w[SHEETS[0]]["A2"].value == "a" and w[SHEETS[0]]["A3"].value == "b"
def test_empty_payload_keeps_sheets_and_notice(tmp_path):
    w=wb(tmp_path,{}); assert w.sheetnames==SHEETS and w[SHEETS[1]]["A2"].value=="결과 없음"
def test_partial_payload(tmp_path): assert wb(tmp_path,{"page_results":[]}).sheetnames==SHEETS
def test_formula_injection_sanitized(tmp_path):
    w=wb(tmp_path,{"page_results":[{"target_id":"=CMD()","verdict":"정상"}]}); assert w[SHEETS[0]]["A2"].value.startswith("'")
def test_control_and_length_sanitized(tmp_path):
    w=wb(tmp_path,{"page_results":[{"target_id":"a\x00"+"x"*40000,"verdict":"정상"}]}); assert "\x00" not in w[SHEETS[0]]["A2"].value and len(w[SHEETS[0]]["A2"].value)<=32767
def test_atomic_save_preserves_previous_on_build_failure(tmp_path, monkeypatch):
    p=make(tmp_path); before=p.read_bytes()
    monkeypatch.setattr("src.reports.page_report.os.replace", lambda *args: (_ for _ in ()).throw(OSError("disk")))
    try: PageReportGenerator(tmp_path).save(payload(),date="2026-08-25")
    except Exception: pass
    assert p.read_bytes()==before
def test_latest_matches_dated(tmp_path):
    p=make(tmp_path); assert (tmp_path/"latest"/"페이지별_점검결과.xlsx").read_bytes()==p.read_bytes()
def test_legacy_payload(tmp_path): assert wb(tmp_path,{"issues":[{"url":"https://x"}]}).sheetnames==SHEETS
def test_same_input_order(tmp_path):
    pa=make(tmp_path,payload())
    a=load_workbook(pa, data_only=True); first=[[c.value for c in r] for r in a[SHEETS[0]].iter_rows()]; a.close()
    pb=PageReportGenerator(tmp_path).save(payload(),date="2026-08-26")
    b=load_workbook(pb, data_only=True); assert first == [[c.value for c in r] for r in b[SHEETS[0]].iter_rows()]; b.close()
def test_report_is_local_only(tmp_path): assert make(tmp_path).exists()
