import json
from openpyxl import load_workbook
from src.daily_pipeline import DailyPipeline
from src.quality.aggregation import issue_key
from src.reports.page_report import SHEETS
BASE="https://fixture.test"
def fx(n):
    links=["/a","/c","/b"]+(["/d"] if n>=2 else []); xml="<urlset xmlns='http://www.sitemaps.org/schemas/sitemap/0.9'>"+"".join(f"<url><loc>{BASE}{x}</loc></url>" for x in links)+"</urlset>"
    p={BASE+"/":{"status_code":200,"html":"<body>"+"".join(f"<a href='{x}'>x</a>" for x in links)+"</body>"},BASE+"/sitemap.xml":{"status_code":200,"text":xml},BASE+"/a":{"status_code":404,"html":"<body>A</body>"},BASE+"/c":{"status_code":410 if n>=2 else 404,"html":"<body>C</body>"},BASE+"/b":{"status_code":200,"html":"<body><img src='/pic.png'"+(" alt='ok'" if n==2 else "")+"></body>"},BASE+"/d":{"status_code":404 if n==2 else 200,"html":"<body>D</body>"}}
    return {"menu":{"main_selectors":["a"],"all_menu_selectors":["a"],"sitemap_path":"/sitemap.xml"},"responses":p}
def run3(t):
    p=DailyPipeline(state_dir=t/"s",output_root=t/"o"); out=[]
    for n,d in enumerate(("2026-08-25","2026-08-26","2026-08-27"),1): out.append(p.run_raw_fixture(fx(n),run_id=f"r{n}",date=d))
    return out
def st(t): return json.loads((t/"s"/"issues.json").read_text(encoding="utf-8"))
def test_a_new_persistent_persistent(tmp_path):
    run3(tmp_path); k=issue_key("fixture",BASE+"/a","connectivity",BASE+"/a","http_status"); assert any(x["issue_key"]==k for x in st(tmp_path)["active_issues"])
def test_b_new_resolved_recurred(tmp_path):
    run3(tmp_path); x=[x for x in st(tmp_path)["active_issues"] if x.get("page_url")==BASE+"/b" and x.get("check_code")=="image-alt"]; assert x and x[0]["recurred_count"]>=1 and len(x[0]["lifecycle_history"])==3
def test_c_fingerprint_change(tmp_path):
    run3(tmp_path); k=issue_key("fixture",BASE+"/c","connectivity",BASE+"/c","http_status"); x=[x for x in st(tmp_path)["active_issues"] if x["issue_key"]==k][0]; assert x["previous_fingerprint"] and "reason" in x["changed_fields"]
def test_d_resolved_history(tmp_path):
    run3(tmp_path); k=issue_key("fixture",BASE+"/d","connectivity",BASE+"/d","http_status"); assert any(x["issue_key"]==k and x.get("resolved_at") for x in st(tmp_path)["resolved_issues"])
def test_json_xlsx_keys_and_dates(tmp_path):
    out=run3(tmp_path); s=st(tmp_path); w=load_workbook(out[-1]["report_path"],data_only=True); text=" ".join(str(v) for row in w[SHEETS[5]].iter_rows(values_only=True) for v in row if v); assert any(x["issue_key"] in text for x in s["active_issues"]+s["resolved_issues"])
def test_fields_and_history_count(tmp_path):
    run3(tmp_path); s=st(tmp_path); assert len((tmp_path/"s"/"run_history.jsonl").read_text().splitlines())==3; assert all(x.get("page_url") and x.get("check_code") and x.get("fingerprint") for x in s["active_issues"]+s["resolved_issues"])
